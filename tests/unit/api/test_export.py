"""CSV and Excel export, against the real read-only engine.

The researcher's real ask is a spreadsheet. These prove it streams, carries a BOM
(so Excel opens accented text right), and cannot be used to escape read-only.
"""

from __future__ import annotations

import csv
import io

import pytest

pytestmark = pytest.mark.integration


def test_csv_has_a_bom_and_header(replica):
    from lake.api import export

    data = b"".join(export.stream_csv("SELECT series_code, year FROM lake.observations LIMIT 3"))
    assert data.startswith(b"\xef\xbb\xbf")  # BOM for Excel

    text = data.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    assert rows[0] == ["series_code", "year"]
    assert len(rows) == 4  # header + 3


def test_csv_preserves_nulls_as_empty(replica):
    from lake.api import export

    data = b"".join(export.stream_csv("SELECT series_code, value FROM lake.observations"))
    text = data.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    deu = next(r for r in rows if r[0] == "DEU")
    assert deu[1] == ""  # NULL gdp -> empty cell, not "None"


def test_xlsx_is_a_valid_workbook(replica):
    from openpyxl import load_workbook

    from lake.api import export

    content = export.build_xlsx("SELECT series_code, year, value FROM lake.observations LIMIT 5")
    assert content[:2] == b"PK"  # xlsx is a zip

    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["series_code", "year", "value"]
    assert ws.max_row == 6  # header + 5


def test_export_cannot_run_a_write(replica):
    """Even bypassing the route validator, the engine refuses a write."""
    from lake.api import export

    with pytest.raises(Exception):  # noqa: B017 — InvalidInputException from DuckDB
        list(export.stream_csv("DELETE FROM lake.observations"))


def test_safe_filename_strips_dangerous_characters():
    from lake.api.export import safe_filename

    assert safe_filename("gdp_annual", "csv") == "gdp_annual.csv"
    assert safe_filename("../../etc/passwd", "csv") == "etc_passwd.csv"
    assert safe_filename("", "xlsx") == "export.xlsx"


# -- one resource, three representations --------------------------------------
#
# `/rows` is the rows. JSON, CSV and Excel are three ways of writing them down, and
# which one you get is a property of the request rather than of the path — a `.csv` on
# the end of a URL is a filename pretending to be a resource.


def test_the_default_representation_is_json(client):
    r = client.get("/api/data/observations/rows")

    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/json")
    assert r.json()["total"] == 10


def test_accept_text_csv_negotiates_a_csv(client):
    """The REST mechanism, for the clients that can use it."""
    r = client.get("/api/data/observations/rows", headers={"Accept": "text/csv"})

    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    assert r.content.startswith(b"\xef\xbb\xbf")  # BOM, for Excel


def test_a_client_that_cannot_send_accept_still_gets_its_csv(client):
    """The reason `?format=` exists, and the regression that motivated the whole design.

    `pandas.read_csv(url)` sends **no Accept header at all**. Under Accept-only
    negotiation it receives JSON — and `read_csv` parses JSON as CSV *without raising*,
    handing the reader an empty DataFrame whose one column name is a blob of JSON. No
    error, no reason why. A browser's `<a href>` download and R's `read.csv` fail the
    same silent way.

    So: no Accept and no param is JSON (right — that is what an unadorned GET should
    be), and the param is what a one-liner uses to say otherwise.
    """
    plain = client.get("/api/data/observations/rows")  # what pandas actually sends
    assert plain.headers["content-type"].startswith("application/json")

    asked = client.get("/api/data/observations/rows", params={"format": "csv"})
    assert asked.headers["content-type"].startswith("text/csv")
    assert asked.content.startswith(b"\xef\xbb\xbf")


def test_the_format_param_beats_a_conflicting_accept_header(client):
    """A caller who typed `?format=csv` was being explicit. A browser sends
    `Accept: text/html,...` on a download link and means nothing by it."""
    r = client.get(
        "/api/data/observations/rows",
        params={"format": "csv"},
        headers={"Accept": "application/json"},
    )

    assert r.headers["content-type"].startswith("text/csv")


def test_every_representation_says_it_varies_by_accept(client):
    """Without `Vary: Accept`, a cache in front of the API can hand a CSV body to the
    next client that asked for JSON."""
    for params in ({}, {"format": "csv"}, {"format": "xlsx"}):
        r = client.get("/api/data/observations/rows", params=params)
        assert r.headers.get("vary") == "Accept", params


def test_an_unknown_format_is_rejected_by_name(client):
    r = client.get("/api/data/observations/rows", params={"format": "parquet"})

    assert r.status_code == 422
    assert "unknown format" in r.json()["detail"]
    assert "csv" in r.json()["detail"]  # it names the ones that do exist


def test_format_is_a_control_not_a_column_filter(client):
    """`format` is reserved. Without that it would be read as a filter on a column
    called `format`, and 422 as an unknown column."""
    assert client.get("/api/data/observations/rows", params={"format": "json"}).status_code == 200


# -- the file -----------------------------------------------------------------


def test_xlsx_is_a_real_workbook(client):
    r = client.get("/api/data/observations/rows", params={"format": "xlsx"})

    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"


def test_a_file_is_named_after_the_thing(client):
    """The URL has lost its extension; the *file* keeps one, because a file on disk
    should say what it is. And it is named after what it holds."""
    r = client.get("/api/data/observations/rows", params={"format": "csv"})
    assert 'filename="observations.csv"' in r.headers["content-disposition"]


def test_a_file_can_be_named_by_the_caller(client):
    r = client.get(
        "/api/data/observations/rows",
        params={"format": "csv", "select": "series_code", "filename": "custom"},
    )
    assert r.status_code == 200
    assert 'filename="custom.csv"' in r.headers["content-disposition"]


def test_downloading_a_series_by_id_gives_that_series_named_after_itself(client):
    """The whole argument for the id. A reader downloading one series wants its rows and
    wants `United_States.csv` in their downloads folder — not the whole table, and not a
    fourth file called `observations.csv`."""
    from lake.api import catalog

    usa = catalog.id_for("gdp_annual", "NY.GDP.MKTP.CD", "United States")
    r = client.get(f"/api/data/{usa}/rows", params={"format": "csv", "select": "series,year"})

    assert r.status_code == 200
    assert 'filename="United_States.csv"' in r.headers["content-disposition"]

    lines = r.content.decode("utf-8-sig").strip().splitlines()
    assert lines[0] == "series,year"
    assert {line.split(",")[0] for line in lines[1:]} == {"United States"}
    assert len(lines) == 3  # header + the two USA rows, not the other eight


# -- what a limit means depends on what you asked for -------------------------


def test_a_file_defaults_to_everything_and_a_page_defaults_to_a_page(client):
    """A page is a screen; a file is the data. A CSV that silently stopped at the JSON
    page size would be the same class of bug as the pandas one above — quiet, and wrong.
    """
    page = client.get("/api/data/observations/rows").json()
    assert page["row_count"] == 10  # the fixture is small; the point is it pages
    assert page["limit"] == 1000  # DEFAULT_LIMIT, not EXPORT_MAX_ROWS

    csv_rows = (
        client.get("/api/data/observations/rows", params={"format": "csv"})
        .content.decode("utf-8-sig")
        .strip()
        .splitlines()[1:]
    )
    assert len(csv_rows) == 10  # everything that matched, with no limit asked for


def test_an_explicit_limit_is_honoured_in_either_representation(client):
    """It is a *default*, not a rule. A caller who says `?limit=3` means it."""
    page = client.get("/api/data/observations/rows", params={"limit": 3}).json()
    assert page["row_count"] == 3
    assert page["has_more"] is True

    csv_rows = (
        client.get("/api/data/observations/rows", params={"format": "csv", "limit": 3})
        .content.decode("utf-8-sig")
        .strip()
        .splitlines()[1:]
    )
    assert len(csv_rows) == 3


def test_a_nonsense_limit_is_rejected(client):
    assert client.get("/api/data/observations/rows", params={"limit": "many"}).status_code == 422
    assert client.get("/api/data/observations/rows", params={"limit": 0}).status_code == 422
    assert client.get("/api/data/observations/rows", params={"offset": -1}).status_code == 422


# -- a file is the same read as the page it was downloaded from ---------------


def test_a_file_carries_the_same_filters_as_the_page(client):
    """An export that filtered differently from the rows it was downloaded from would be
    a quiet, undetectable lie. They share one `_compile_rows`, so they cannot drift."""
    r = client.get(
        "/api/data/observations/rows",
        params={"format": "csv", "series_code": "USA", "select": "series_code,year"},
    )
    lines = r.content.decode("utf-8-sig").strip().splitlines()

    assert lines[0] == "series_code,year"
    assert {line.split(",")[0] for line in lines[1:]} == {"USA"}
    assert len(lines) == 3  # header + the two USA rows


def test_an_injected_filter_on_a_file_finds_nothing_and_writes_nothing(client, tmp_path):
    target = tmp_path / "exfil.csv"
    r = client.get(
        "/api/data/observations/rows",
        params={"format": "csv", "series_code": f"x'; COPY (SELECT 1) TO '{target}'; --"},
    )

    assert r.status_code == 200
    # Header only: it is a value, so it matched no row. And it never became SQL.
    assert r.content.decode("utf-8-sig").strip().count("\n") == 0
    assert not target.exists()


def test_a_file_naming_no_column_is_rejected(client):
    r = client.get("/api/data/observations/rows", params={"format": "csv", "nope": "1"})
    assert r.status_code == 422


def test_a_file_of_an_unknown_id_is_404(client):
    r = client.get("/api/data/zzzzzzzz/rows", params={"format": "csv"})
    assert r.status_code == 404
