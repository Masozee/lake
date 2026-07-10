"""Pure function: xlsx bytes -> list[dict]. No network, no disk.

TEMPLATE. The column layout below is a plausible guess, not a verified one. Before
you trust it:

    1. run `lake scrape bps_inflation` once, so real bytes land in raw/
    2. copy one file into tests/fixtures/bps_inflation_sample.xlsx
    3. open it, fix the constants below, and write a test against the fixture

A parser that looks authoritative but was never run against real bytes is worse
than no parser: it will silently produce plausible nonsense. Nothing else in the
pipeline imports this module until you wire it into transform.py, so an
unfinished parser cannot corrupt raw/.

Needs the `transform` extra (openpyxl). The scraper does not import this, so
`lake scrape bps_inflation` works without it.
"""

from __future__ import annotations

import io
import re
from datetime import date
from typing import Any

# Indonesian and English month names, lowercased. Extend as needed.
MONTHS = {
    "januari": 1,
    "january": 1,
    "februari": 2,
    "february": 2,
    "maret": 3,
    "march": 3,
    "april": 4,
    "mei": 5,
    "may": 5,
    "juni": 6,
    "june": 6,
    "juli": 7,
    "july": 7,
    "agustus": 8,
    "august": 8,
    "september": 9,
    "oktober": 10,
    "october": 10,
    "november": 11,
    "desember": 12,
    "december": 12,
}

#: 0-based column holding the region name. Verify against a real file.
REGION_COL = 0
#: 1-based row holding the month names. Verify against a real file.
HEADER_ROW = 1


def month_number(label: str) -> int | None:
    return MONTHS.get(label.strip().lower())


def parse(raw: bytes, *, sheet: str | None = None) -> list[dict[str, Any]]:
    """Rows of (region, period, inflation_pct).

    A blank cell is a missing observation, not a zero. A region with no survey
    that month is not a region with zero inflation, and coercing one into the
    other is how a dataset quietly becomes wrong.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("bps_inflation parsing needs: uv sync --extra transform") from exc

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.worksheets[0]
        rows = worksheet.iter_rows(min_row=HEADER_ROW, values_only=True)

        header = next(rows, None)
        if not header:
            return []

        months = {
            i: m
            for i, cell in enumerate(header)
            if isinstance(cell, str) and (m := month_number(cell)) is not None
        }
        if not months:
            raise ValueError(f"no month columns in header row {HEADER_ROW}: {header!r}")

        year = infer_year(worksheet.title) or infer_year(str(header))
        if year is None:
            raise ValueError(f"could not determine the year from sheet {worksheet.title!r}")

        records: list[dict[str, Any]] = []
        for row in rows:
            region = row[REGION_COL] if len(row) > REGION_COL else None
            if not isinstance(region, str) or not region.strip():
                continue

            for col, month in months.items():
                value = row[col] if col < len(row) else None
                if value is None or value == "":
                    continue  # missing observation — do NOT coerce to 0.0
                try:
                    pct = float(value)
                except (TypeError, ValueError):
                    continue
                records.append(
                    {"region": region.strip(), "period": date(year, month, 1), "inflation_pct": pct}
                )
        return records
    finally:
        workbook.close()


def infer_year(text: str) -> int | None:
    """The year lives in the sheet title or a banner cell, never in a column."""
    match = re.search(r"(?:19|20)\d{2}", text)
    return int(match.group(0)) if match else None
